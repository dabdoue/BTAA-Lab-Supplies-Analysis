# %%
'''
whenever you see the above symbol (# %%) it means that all the code in between that line of code
and the next occurence of that set of characters is all considered to be one "cell"
a cell is a section of code that can be run independently of the rest of the code
splitting the code up into cells allows us to not run parts of the code over and over again 
while testing it

This is especially useful as this first cell takes about 1 minute to run on my laptop.
basically, this means you only need to run this cell once. 

What this cell does is read in the excel files: one with fisher's claimed data, the other with
our actual purchasing data, and stores all the data into variables that the code can then go through
'''

import re
import numpy as np
from collections import defaultdict
import pandas as pd
import os
from openpyxl import load_workbook
import traceback


def fisher_school_price_comparison(year):
    try:
        '''
        what the following code does is get the files that have been uploaded to the uploads folder,
        looks at their names, and for the file that was uploaded containing "master" somewhere in the name,
        it sets that file as the reference for the master data. 

        the same is done for the fisher proposed file. if the word "fisher" is present in the file name,
        it sets that file as the reference for the fisher data. 

        it then gets the first sheet in that file as the sheet name to reference for both files
        '''
        # makes default data analysis year 2022 if no year given
        if year == "":
            year = "2022"
        
        if "uploads" not in os.getcwd():
            os.chdir(os.path.join(os.getcwd(), "uploads"))
        data_files = os.listdir(os.getcwd())

        master_data_file = ""
        master_sheet_name = ""
        fisher_data_file = ""
        fisher_sheet_name = ""

        for data_file in data_files:
            if "master" in data_file.lower():
                master_data_file = data_file
                master_sheet_name = load_workbook(
                data_file, read_only=True, keep_links=False).sheetnames[0]
            elif "fisher" in data_file.lower():
                fisher_data_file = data_file
                fisher_sheet_name = load_workbook(
                    data_file, read_only=True, keep_links=False).sheetnames[0]

        school_rows = pd.read_excel(master_data_file, sheet_name=master_sheet_name).to_numpy()
        fisher_rows = pd.read_excel(fisher_data_file, sheet_name=fisher_sheet_name).to_numpy()

        # %%

        '''
        what this cell does is sort through all the data from the fisher price data file, clean up the catalog numbers,
        and sort everything based on catalog number. The reason sorting is important is because if the 
        data is not sorted, then it will take a very very long time to run. Sorting data is very common
        when working very large sets of data like this
        '''

        # this line just gets rid of the first row of the data, since its just the column titles
        fisher_rows = fisher_rows[1:]

        # the purpose of everything in this for loop is to remove any leading zeroes from the catalog numbers,
        # as well as remove any commas from the data as commas lead to issues exporting the data later
        for idx in range(len(fisher_rows)):
            catalog_num = str(fisher_rows[idx][0])

            # this if statement checks to see if there is a space in the catalog number, and if there is, only
            # consider the characters before the space to be the catalog number that we want
            # this handles cases where the catalog number is something like "123456 (EA)"
            if " " in catalog_num:
                fisher_rows[idx][0] = catalog_num[:catalog_num.index(" ")]

            fisher_rows[idx][0] = catalog_num.lstrip("0") # this gets rid of leading zeroes
            fisher_rows[idx][1] = fisher_rows[idx][1].replace(',', '') # this removes commas from the data

        sorted_fisher_rows = sorted(fisher_rows, key=lambda row: str(row[0])) # this sorts the data based on catalog number alphabetically not by size!

        # %%
        '''
        this cell has the same purpose of the above cell, to clean up the data from the excel file,
        but it does so for the data from all the schools actual purchasing data

        here there is more handling of weird catalog numbers, and then sorting the data
        '''
        schools = []
        sorted_school_rows = []
        for idx in range(len(school_rows)):
            # if the year of the data is not desired year range, then skip it
            if str(school_rows[idx][1]) not in year:
                continue
            
            # # This code will make it so that only items labeled as "FISHER SCIENTIFIC" in the master file
            # # will be used for analysis
            # if str(school_rows[idx][5]) != "FISHER SCIENTIFIC":
            #     continue

            
            catalog_num = str(school_rows[idx][6])

            # if the catalog number is empty, skip this piece of data
            if catalog_num == '':
                continue
            
            # if the "|" character is in the catalog number, only consider the catalog number to be what is after the "|"
            # for some reason a lot of Penn State's catalog numbers have this
            if "|" in catalog_num:
                catalog_num = catalog_num[catalog_num.index("|") + 1:]

            # if space in catalog num, only take characters before space as the catalog num
            if " " in catalog_num:
                catalog_num = catalog_num[:catalog_num.index(
                    " ")]

            # we are keeping track of all the different schools in the data, so if the current school's data is not 
            # in the list of schools thus far, update the lsit
            if school_rows[idx][3] not in schools:
                schools.append(school_rows[idx][3])

            school_rows[idx][6] = catalog_num.lstrip("0") # this gets rid of leading zeroes
            sorted_school_rows.append(school_rows[idx])

        # this sorts the data based on catalog number first alphabetically not by size, and if tied, by the school name
        # print(len(sorted_school_rows))
        if len(sorted_school_rows) == 0:
            return "No data found for year: " + year

        sorted_school_rows = sorted(
            sorted_school_rows, key=lambda row1: [row1[6], row1[3]])  

        # %%
        '''
        this cell creates the titles for all of the data that will be output later. it does so based on the "schools" variable
        compiled_data is the variable that will be populated with all the data from the schools combined with fisher's data
        '''

        compiled_data = [["Catalog Number", "Product Description", "Fisher Stated Sales",
                        "Fisher Stated List Price", "Fisher Stated Qty (SU + AU)"]]
        num_schools = len(schools)
        for i in range(3 * num_schools):
            compiled_data[0].append("")

        for idx, school in enumerate(schools):
            compiled_data[0][5 + idx] = school + " Cost"
            compiled_data[0][5 + num_schools + idx] = school + " Quantity"
            compiled_data[0][5 + (2 * num_schools) + idx] = school + " Average Cost Per Item"

        compiled_data[0].append("Total Cost")
        compiled_data[0].append("Total Quantity")


        # %%
        '''
        this code's purpose is to further organize the data into a dictionary
        a dictionary is essentially a structure where based on a key, you have value(s) associated with it

        in this scenario, i am taking the first two characters of a catalog number as a key,
        and any other catalog numbers in the data whose first two characters are the same as that key, 
        are then associated as values for that key

        for example:

        lets say you have the following catalog numbers:

        10202, 1234, 1256, 123213, 23546, 236

        you would end up with a dictionary that looks like this:
        {'10': ['10202'], '12': ['1234', '1256', '123213'], '23': ['23546', '236']}

        everything is nicely sorted, so when searching for some data, we first check to see which list of keys we are looking for based
        on the first two characters, and then we can search much more quickly
        '''

        school_catalog_nums = np.array(sorted_school_rows)[:, 6].astype('str') # this puts all the catalog nums from the data into a single array
        catalog_nums_dict = defaultdict(list)
        for catalog_num in school_catalog_nums:
            if catalog_num != '':
                catalog_nums_dict[catalog_num[:2]].append(catalog_num) # this does the dictionary sorting as mentioned in the description above


        # %%
        '''
        this code's overall purpose is to match all the items in fisher's file to the items in our records

        it goes through each item in the fisher file, cleans up some data, prepares the list which will 
        contain all the aggregated data, and then searches dictionary that we created earlier, and then 
        populates the list with all the results from our own data.

        This also accounts for if there are multiple instances of the same catalog number in our data,
        including whether the instances are from different schools or from the same school. If from the same
        school, it sums the quantities and costs together.
        '''
        i = 0
        last_school_idx = 0
        matched_data = []
        for fr in sorted_fisher_rows: # going through each item in the fisher data file
            i += 1
            if i <= 1:
                continue
            catalog_num = str(fr[0])
            su_num = re.sub(r'[^0-9.]', '', str(fr[4])) # ensures that the data only consists of numbers and .
            au_num = re.sub(r'[^0-9.]', '', str(fr[5]))
            if su_num == "": # sometimes values are blank, so we assign values of 0 to them
                su_num = 0
            if au_num == "":
                au_num = 0

            cur_data = [fr[0], fr[1], re.sub(r'[^0-9.]', '', str(fr[2])), re.sub(
                r'[^0-9.]', '', str(fr[3])), float(su_num) + float(au_num)] # initializes the data for the current item with values from fisher's file

            matched_idxs = []
            for matching_idx, num in enumerate(catalog_nums_dict[catalog_num[:2]]): # checks the dictionary for current catalog num 
                if num == catalog_num:
                    matched_idxs.append(matching_idx) # if the catalog num exists in the dictionary, then save that data

            # below two lines get the keys for all the keys in the dictionary that came before the currently used key
            previous_dicts_keys = list(catalog_nums_dict.keys())
            previous_dicts_keys = previous_dicts_keys[:previous_dicts_keys.index(
                catalog_num[:2])]
            
            # using the length of the value arrays of the previous keys, gets the index of the current catalog number in the full dataset
            prev_idxs_sum = 0
            for key in previous_dicts_keys:
                prev_idxs_sum += len(catalog_nums_dict[key])

            # gets all the actual indexes (not from dictionary) of all the items that have been matched in our dataset
            real_idxs = []
            for cur_matching_idx in matched_idxs:
                real_idxs.append(prev_idxs_sum + cur_matching_idx)

            # itereates through all the indexes we just found
            for matched_idx in real_idxs:

                # assigns the values for the data we are iterating through
                school_name = sorted_school_rows[matched_idx][3]
                school_quantity_text = school_name + " quantity"
                school_quantity = str(sorted_school_rows[matched_idx][11])
                school_cost_text = school_name + " cost"
                school_cost = str(sorted_school_rows[matched_idx][13])

                # if cur_data[5] already contains data, then go into try block
                # essentially checks to see if any previous data on this catalog number has been saved
                try:
                    # if data for this school already exists for this catalog number
                    if school_name in cur_data[5]:
                        school_idx = cur_data[5].index(school_name) # uses universal index of schools so that the data gets placed into correct column
                        cur_data[6][school_idx] = int(cur_data[6][school_idx]) + \
                            int(school_quantity) # adds the num of items to the already existing value
                        # ensures data for cost is correct format
                        school_cost_float = re.sub(
                            r'[^0-9.]', '', school_cost)
                        if school_cost_float == '' or school_cost == 'nan':
                            school_cost_float = 0
                        else:
                            school_cost_float = float(school_cost_float)
                        cur_data[7][school_idx] = float(
                            cur_data[7][school_idx]) + school_cost_float # adds current data's cost to existing cost value

                    # if data exists for this catalog num, but not for this school
                    else:
                        cur_data[5].append(school_name)
                        cur_data[6].append(int(school_quantity))
                        if school_cost == 'nan':
                            school_cost = '0'
                        cur_data[7].append(
                            float(re.sub(r'[^0-9.]', '', school_cost)))
                # if this is the first time the catalog number has any data, populate the list with school name, quantity of items, and total cost of items
                except:
                    total_cost = re.sub(r'[^0-9.]', '', school_cost) # ensures the data only contains numbers and .
                    if total_cost == '':
                        total_cost = 0
                    cur_data.append([school_name])
                    cur_data.append([int(school_quantity)])
                    cur_data.append([float(total_cost)])

            matched_data.append(cur_data) # adds this formated data to the entire list



        # %%
        # this code cell essentially takes all the data that was compiled from the datasets, and organizes it in the format that will show up in the output csv
        num_schools_zeros = list(np.zeros((len(schools) * 3) + 2)) # initializes an empty row of data with the correct amount of zeros
        len_num_zeros = len(num_schools_zeros)
        num_columns = len(compiled_data[0])

        # goes through each item in the data aggregated in the previous cell
        for item in matched_data:
            # makes copy of item in quesiton to 
            item_copy = item[:5]
            item_copy.extend(num_schools_zeros)

            # if there is data in the current data, then add all the data into the list that will be converted to a csv
            if len(item) > 5:
                for idx, school in enumerate(item[5]): # for each school's data, go through
                    if school == '':
                        continue
                    school_idx = schools.index(school)
                    school_quantity = item[6][idx]

                    school_cost = round(float(item[7][idx]), 2) # getting value for cost to 2 decimals

                    # assigning cost and quantity values to list
                    item_copy[school_idx+5] = school_cost 
                    item_copy[school_idx+5+len(schools)] = school_quantity
                    item_copy[school_idx+5+(2*len(schools))
                            ] = round(school_cost / school_quantity, 2) # average cost per item
                    item_copy[num_columns -
                            2] = round(school_cost + item_copy[num_columns - 2], 2) # sum costs from all schools and add as column
                    item_copy[num_columns -
                            1] = int(school_quantity + item_copy[num_columns - 1]) # sum quantities of items from all schools and add as column

            compiled_data.append(item_copy)

        # write compiled data to csv
        f = open('output_file=fisher_price_comparison' + "_" + year + '.csv', 'w', encoding="utf-8") 
        for item in compiled_data:
            f.write(','.join([str(x) for x in item]) + '\n')
        f.close()

        return "True"
    except Exception:
        return traceback.print_exc()




