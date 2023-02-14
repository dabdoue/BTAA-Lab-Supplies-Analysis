import collections
import csv
import re
import pandas as pd
import os
from openpyxl import load_workbook
import math
import traceback

def tracking_market_baskets():
    try:
        i = 0
        prices_2019 = {}
        prices_2020 = {}
        prices_2021 = {}
        prices_2022 = {}

        desc_2019 = {}
        desc_2020 = {}
        desc_2021 = {}
        desc_2022 = {}
        all_prices_new = collections.defaultdict(list)
        all_prices_old = collections.defaultdict(list)
        # make variables for each year, have it be dictionary that is catalog number and price


        def find_matching_item(idx, prices, descs, key):
            item_desc = list(desc_2022.values())[idx]
            try:
                catalog_num = list(prices)[list(
                    descs.values()).index(item_desc)]
                price = prices[catalog_num]
                all_prices_new[key.lstrip('0')].append(
                    str(price) + " (matched with similar item: " + str(catalog_num) + ")")
            except:
                all_prices_new[key.lstrip('0')].append("")

        if "uploads" not in os.getcwd():
            os.chdir(os.path.join(os.getcwd(), "uploads"))
        data_file = os.listdir(os.getcwd())[0]
        sheet_name = load_workbook(data_file, read_only=True,
                                keep_links=False).sheetnames[0]
        rows = pd.read_excel(
            data_file, sheet_name=sheet_name).to_numpy()

        for r in rows:
            
            if isinstance(r[0], float) and not math.isnan(r[0]):
                r[0] = int(r[0])
            r = [str(i) for i in r]
            for i in range(len(r)):
                if r[i] == 'nan':
                    r[i] = ''
                    
            if r[19] == "2019":
                prices_2019[r[2]] = r[8]
                desc_2019[r[2]] = r[3]

            if r[19] == "2020":
                prices_2020[r[2]] = r[8]
                desc_2020[r[2]] = r[3]

            if r[19] == "2021":
                prices_2021[r[2]] = r[8]
                desc_2021[r[2]] = r[3]

            if r[19] == "2022":
                prices_2022[r[2]] = r[8]
                desc_2022[r[2]] = r[3]

        # print(list(prices_2019.keys())[0])
        # print(list(prices_2020.values())[0])
        # print(list(prices_2021.values())[0])
        # print(list(prices_2022.values())[0])

        for idx, key in enumerate(prices_2022):
            try:
                all_prices_new[key.lstrip('0')].append(prices_2022[key])
                all_prices_old[key.lstrip('0')].append(prices_2022[key])

            except:
                all_prices_new[key.lstrip('0')] = prices_2022[key]
                all_prices_old[key.lstrip('0')] = prices_2022[key]

            if key in prices_2021:
                all_prices_new[key.lstrip('0')].append(prices_2021[key])
            else:
                find_matching_item(idx, prices_2021, desc_2021, key)
            if key in prices_2020:
                all_prices_new[key.lstrip('0')].append(prices_2020[key])
            else:
                find_matching_item(idx, prices_2020, desc_2020, key)
            if key in prices_2019:
                all_prices_new[key.lstrip('0')].append(prices_2019[key])
            else:
                find_matching_item(idx, prices_2019, desc_2019, key)

        # print(all_prices_new)

        descs_2022_list = list(desc_2022.values())

        with open('output_file=all_in_market_basket.csv', 'w') as f:
            with open('output_file=consistent_items.csv', 'w') as g:
                g.write("%s, %s, %s, %s, %s, %s\n" % ("Catalog Number",
                        "Description", "2022", "2021", "2020", "2019"))
                f.write("%s, %s, %s, %s, %s, %s\n" % ("Catalog Number",
                        "Description", "2022", "2021", "2020", "2019"))
                for idx, key in enumerate(all_prices_new.keys()):
                    if all_prices_new[key][0] != "" and all_prices_new[key][1] != "" and all_prices_new[key][2] != "" and all_prices_new[key][3] != "":
                        g.write("%s, %s, %s, %s, %s, %s\n" % (
                            key, ''.join(descs_2022_list[idx].split(',')), all_prices_new[key][0], all_prices_new[key][1], all_prices_new[key][2], all_prices_new[key][3]))
                    f.write("%s, %s, %s, %s, %s, %s\n" % (
                        key, ''.join(descs_2022_list[idx].split(',')), all_prices_new[key][0], all_prices_new[key][1], all_prices_new[key][2], all_prices_new[key][3]))

        return "True"
    except Exception:
        return traceback.print_exc()
