from cgi import print_exception
import csv
import collections
from re import sub
import numpy as np
import itertools
import os
import pandas as pd
from openpyxl import load_workbook
import traceback

num_top_values = 10

def school_usage_analysis():
    try:
        i = 0
        prices_dict = collections.defaultdict(list)
        volume_dict = collections.defaultdict(list)
        catalog_nums_dict = collections.defaultdict(list)
        descs_dict = collections.defaultdict(list)
        numbered_prices_dict = collections.defaultdict(list)
        numbered_volume_dict = collections.defaultdict(list)

        if "uploads" not in os.getcwd():
            os.chdir(os.path.join(os.getcwd(), "uploads"))
        data_file = os.listdir(os.getcwd())[0]
        sheet_name = load_workbook(data_file, read_only=True,
                                keep_links=False).sheetnames[0]
        rows = pd.read_excel(
            data_file, sheet_name=sheet_name).to_numpy()

        rows = pd.read_excel(
            data_file, sheet_name=sheet_name).to_numpy()

        rows = sorted(rows, key=lambda row: row[3])
        j = 0
        for r in rows:

            r = [str(i) for i in r]
            for i in range(len(r)):
                if r[i] == 'nan':
                    r[i] = ''

            part_number = r[0]
            if r[13] != '':
                prices_dict[r[3]].append(float(sub(r'[^\d.]', '', r[13])))
                volume_dict[r[3]].append(r[11])
                catalog_nums_dict[j].append(r[10])
                descs_dict[j].append(r[8])
                numbered_prices_dict[j] = (float(sub(r'[^\d.]', '', r[13])))
                numbered_volume_dict[j] = int(r[11])

                j += 1

        prices_idx = 0
        f = open('output=school_usage_analysis.csv', 'w', newline='')
        writer = csv.writer(f)

        for idx, school in enumerate(prices_dict):
            prices_arr_size = len(list(prices_dict.values())[idx])

            truncated_prices_dict = dict(itertools.islice(
                numbered_prices_dict.items(), prices_idx, prices_idx + prices_arr_size))
            sorted_tuples = sorted(truncated_prices_dict.items(),
                                key=lambda item: item[1], reverse=True)
            sorted_prices_dict = {k: v for k, v in sorted_tuples}

            truncated_volume_dict = dict(itertools.islice(
                numbered_volume_dict.items(), prices_idx, prices_idx + prices_arr_size))
            sorted_tuples = sorted(truncated_volume_dict.items(),
                                key=lambda item: item[1], reverse=True)
            sorted_volume_dict = {k: v for k, v in sorted_tuples}

            top_by_price_values = list(sorted_prices_dict.values())[:num_top_values]
            top_by_volume_values = list(sorted_volume_dict.values())[:num_top_values]
            top_by_price_keys = list(sorted_prices_dict.keys())[:num_top_values]
            # print(top_by_price_keys)
            top_by_volume_keys = list(sorted_volume_dict.keys())[:num_top_values]
            # print(top_by_volume_keys)

            writer.writerow([school])
            writer.writerow(["", "Top Purchased By Price"])
            for key in top_by_price_keys:
                writer.writerow(["", "", catalog_nums_dict[key][0],
                                ''.join(descs_dict[key]), numbered_prices_dict[key]])
            writer.writerow(["", "Top Purchased By Volume"])
            for key in top_by_volume_keys:
                writer.writerow(["", "", catalog_nums_dict[key][0],
                                ''.join(descs_dict[key]), numbered_volume_dict[key]])

                #     f.write("%s, %s, %s\n" % (str(catalog_nums_dict[key]),
                #                               str(descs_dict[key]), numbered_prices_dict[key]))
            writer.writerow([])
            prices_idx = prices_arr_size + prices_idx

        f.close()
        return "True"
    except Exception:
        return traceback.print_exc()

